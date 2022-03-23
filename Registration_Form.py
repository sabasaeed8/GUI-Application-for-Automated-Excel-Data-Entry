#Name :  SABA
#######################TERM PROJECT################################
#Title : REGISTERATION FORM OR STUDENT INFORMATION SYSTEM
###############################################################################
#In this application, User has to fill up the required information and that 
#information is automatically written into an excel file.
#Firstly, create an empty excel file, after that pass an absolute path of the 
#excel file in the program so that the program is able to access that excel file.
###############################################################################


# import openpyxl and tkinter module
from openpyxl import *
from tkinter import *

# load the existing excel file 
work_book = load_workbook('excel.xlsx')
# create the sheet object
work_sheet = work_book.active


def excel():
    
    # resize the width of columns in excel spreadsheet
    work_sheet.column_dimensions['A'].width = 20
    work_sheet.column_dimensions['B'].width = 30
    work_sheet.column_dimensions['C'].width = 30
    work_sheet.column_dimensions['D'].width = 20
    work_sheet.column_dimensions['E'].width = 10
    work_sheet.column_dimensions['F'].width = 10
    work_sheet.column_dimensions['G'].width = 20
    work_sheet.column_dimensions['H'].width = 40
    work_sheet.column_dimensions['I'].width = 50
    
    
    # write given data to an excel spreadsheet 
	# at particular location 
    work_sheet.cell(1,1).value = 'NAME'
    work_sheet.cell(1,2).value = "FATHER'S NAME"
    work_sheet.cell(1,3).value = 'DEGREE COURSE'
    work_sheet.cell(1,4).value = 'REG.NO'
    work_sheet.cell(1,5).value = 'SEMESTER'
    work_sheet.cell(1,6).value = 'CGPA'
    work_sheet.cell(1,7).value = 'CONTACT NUMBER'
    work_sheet.cell(1,8).value = 'EMAIL ID'
    work_sheet.cell(1,9).value = 'ADDRESS'
    
    
# Function to set focus (cursor) 
def focus1(event):
    # set focus on the course_field box 
    Father_name_field.focus_set()
 
    
# Function to set focus (cursor) 
def focus2(event):
    # set focus on the course_field box 
    Degree_course_field.focus_set()

    
# Function to set focus (cursor) 
def focus3(event):
    # set focus on the course_field box 
    Reg_number_field.focus_set()  

    
# Function to set focus (cursor) 
def focus4(event):
    # set focus on the course_field box 
    Semester_field.focus_set()   

    
# Function to set focus (cursor) 
def focus5(event):
    # set focus on the course_field box 
    CGPA_field.focus_set()


# Function to set focus (cursor) 
def focus6(event):
    # set focus on the course_field box 
    Contact_number_field.focus_set()    


# Function to set focus (cursor) 
def focus7(event):
    # set focus on the course_field box 
    Email_id_field.focus_set()
    

# Function to set focus (cursor) 
def focus8(event):
    # set focus on the course_field box 
    Address_field.focus_set()
    
    
# Function for clearing the 
# contents of text entry boxes 
def clear(): 
	
	# clear the content of text entry box 
    Name_field.delete(0, END) 
    Father_name_field.delete(0,END)
    Degree_course_field.delete(0, END) 
    Reg_number_field.delete(0,END)
    Semester_field.delete(0, END)
    CGPA_field.delete(0, END) 
    Contact_number_field.delete(0, END)
    Email_id_field.delete(0, END) 
    Address_field.delete(0, END) 
    
    
# Function to notify that user
# is already registered.
def already_registered ():
    
    # if user is already registered it 
    # returns True otherwise False
    r = work_sheet.max_row
    
    for i in range (1,r+1):
        if (Name_field.get().upper() in work_sheet.cell(row=i,column=1).value and 
           Father_name_field.get().upper() in work_sheet.cell(row=i,column=2).value and
           Degree_course_field.get().upper() in work_sheet.cell(row=i,column=3).value and
           Reg_number_field.get() in work_sheet.cell(row=i,column=4).value and
           Semester_field.get() in work_sheet.cell(row=i,column=5).value and
           CGPA_field.get() in work_sheet.cell(row=i,column=6).value and
           Contact_number_field.get() in work_sheet.cell(row=i,column=7).value and  
		   Email_id_field.get().lower() in work_sheet.cell(row=i,column=8).value and
		   Address_field.get().lower() in work_sheet.cell(row=i,column=9).value) :
            
            return True
        
    return False
            
             
# Function to take data from GUI 
# window and write to an excel file 
def insert () :
    
    # if user did not fill any entry 
	# then print "empty input. please enter your detail." 
    if (Name_field.get().upper() == "" and
		Father_name_field.get().upper() == "" and
		Degree_course_field.get().upper() == "" and
		Reg_number_field.get() == "" and
        Semester_field.get() == "" and
        CGPA_field.get() == "" and
        Contact_number_field.get() == "" and  
		Email_id_field.get().lower() == "" and
		Address_field.get().lower() == "") :
        
        print('EMPTY INPUT. PLEASE ENTER YOUR DETAIL.')
        
    
    # if user missed  any entry
    # then print please enter your complete detail.
    elif (Name_field.get().upper() == "" or
		Father_name_field.get().upper() == "" or
		Degree_course_field.get().upper() == "" or
		Reg_number_field.get() == "" or
        Semester_field.get() == "" or
        CGPA_field.get() == "" or
        Contact_number_field.get() == "" or 
		Email_id_field.get().lower() == "" or
		Address_field.get().lower() == "") :
        
        print('PLEASE ENTER YOUR COMPLETE DETAIL.')

        
    # if user is already registered  
    # then print This user is already registered.
    elif already_registered():
        print('THIS USER IS ALREADY REGISTERED.')
        clear()
        
        
    else:
        
        # assigning the max row and max column 
		# value upto which data is written 
		# in an excel sheet to the variable 
        current_row = work_sheet.max_row
        current_column = work_sheet.max_column
    
        # get method returns current text 
		# as string which we write into 
		# excel spreadsheet at particular location
        work_sheet.cell(current_row+1 , column=1).value = Name_field.get().upper()
        work_sheet.cell(current_row+1 , column=2).value = Father_name_field.get().upper()
        work_sheet.cell(current_row+1 , column=3).value = Degree_course_field.get().upper()
        work_sheet.cell(current_row+1 , column=4).value = Reg_number_field.get()
        work_sheet.cell(current_row+1 , column=5).value = Semester_field.get()
        work_sheet.cell(current_row+1 , column=6).value = CGPA_field.get()
        work_sheet.cell(current_row+1 , column=7).value = Contact_number_field.get()
        work_sheet.cell(current_row+1 , column=8).value = Email_id_field.get().lower()
        work_sheet.cell(current_row+1 , column=9).value = Address_field.get().lower()
        
        # save the file
        work_book.save('excel.xlsx')
        
        # set focus on the name_field box
        Name_field.focus_set()
        
        # call the clear function
        clear()
        print('SUCCESSFULLY REGISTERED!!!')


# Driver code 
if __name__ == "__main__": 
    
    # create a GUI window
    root = Tk()
    
    # set the title of GUI window
    root.title('Registeration Form')
    
    # set the background colour of GUI window
    root.configure(bg = 'grey')
    
    # set the geometry of GUI window
    root.geometry('500x550')
    
    excel()
    
    # create a form heading
    heading = Label(root, text='Registeration Form', font=('arial',20,'bold'), bg='black', fg='white')
    heading.grid(row=0,column=0)
    
    # create a name label
    Name = Label(root, text='Name', bg='grey', fg='white', font=('arial',15,'bold') ) 
    Name.grid(row=1,column=0)
    
    # create a Father name label
    Father_name = Label(root, text="Father's Name", bg='grey', fg='white', font=('arial',15,'bold') )
    Father_name.grid(row=3,column=0)
    
    # create a degree course label
    Degree_course = Label(root, text='Degree Course', bg='grey', fg='white', font=('arial',15,'bold') )
    Degree_course.grid(row=5,column=0)
    
    #create a Reg number label
    Reg_number = Label(root, text="Registeration Number", bg='grey', fg='white', font=('arial',15,'bold') )
    Reg_number.grid(row=7,column=0)
    
    # create a semester label
    Semester = Label(root, text="Semester", bg='grey', fg='white', font=('arial',15,'bold') )
    Semester.grid(row=9,column=0)
    
    # create a CGPA label
    CGPA = Label(root, text="CGPA", bg='grey', fg='white', font=('arial',15,'bold') )
    CGPA.grid(row=11,column=0)
    
    # create a contact number label
    Contact_number = Label(root, text="Contact Number", bg='grey', fg='white', font=('arial',15,'bold') )
    Contact_number.grid(row=13,column=0)
    
    # create a email id label
    Email_id = Label(root, text="Email id", bg='grey', fg='white', font=('arial',15,'bold') )
    Email_id.grid(row=15,column=0)
     
    # create a address label
    Address = Label(root, text="Address", bg='grey', fg='white', font=('arial',15,'bold') )
    Address.grid(row=17,column=0)

    
    # create a text entry box 
	# for typing the information 
    Name_field = Entry(root)
    Name_field.grid(row=2,column=0,ipadx='150')
    
    Father_name_field = Entry(root)
    Father_name_field.grid(row=4,column=0,ipadx='150')
    
    Degree_course_field = Entry(root)
    Degree_course_field.grid(row=6,column=0,ipadx='150')
    
    Reg_number_field = Entry(root)
    Reg_number_field.grid(row=8,column=0,ipadx='150')
    
    Semester_field = Entry(root)
    Semester_field.grid(row=10,column=0,ipadx='150')
    
    CGPA_field = Entry(root)
    CGPA_field.grid(row=12,column=0,ipadx='150')
    
    Contact_number_field = Entry(root)
    Contact_number_field.grid(row=14,column=0,ipadx='150')
    
    Email_id_field = Entry(root)
    Email_id_field.grid(row=16,column=0,ipadx='150')
    
    Address_field = Entry(root)
    Address_field.grid(row=18,column=0,ipadx='150')
    
    
    # bind method of widget is used for 
	# the binding the function with the events 

	# whenever the enter key is pressed 
	# then call the focus1 function
    Name_field.bind("<Return>", focus1)
    
    # whenever the enter key is pressed 
	# then call the focus2 function
    Father_name_field.bind("<Return>", focus2)
    
    # whenever the enter key is pressed 
	# then call the focus3 function
    Degree_course_field.bind("<Return>", focus3)
    
    # whenever the enter key is pressed 
	# then call the focus4 function
    Reg_number_field.bind("<Return>", focus4)
    
    # whenever the enter key is pressed 
	# then call the focus5 function
    Semester_field.bind("<Return>", focus5)
    
    # whenever the enter key is pressed 
	# then call the focus6 function
    CGPA_field.bind("<Return>", focus6)
    
    # whenever the enter key is pressed 
	# then call the focus7 function
    Contact_number_field.bind("<Return>", focus7)
    
    # whenever the enter key is pressed 
	# then call the focus8 function
    Email_id_field.bind("<Return>", focus8)
    
    # call excel function 
    excel()

	# create a Submit Button and place into the root window 
    submit = Button(root, text="Submit",font=('arial',15,'bold'), fg="white", bg="black", command=insert) 
    submit.grid(row=19,column=0)
    

	# start the GUI 
    root.mainloop()
    
###############################################################################